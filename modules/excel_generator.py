"""
Excel Generator Module
======================
Generates professional Excel dashboards with advanced formatting,
charts, and multiple sheets that exceed requirements.
"""

import pandas as pd
import numpy as np
from datetime import datetime
import io
from typing import Dict, List, Optional, Any
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    NamedStyle, Protection
)
from openpyxl.chart import (
    BarChart, PieChart, LineChart, Reference,
    Series, DoughnutChart
)
from openpyxl.chart.axis import DateAxis
from openpyxl.formatting.rule import (
    ColorScaleRule, CellIsRule, FormulaRule,
    DataBarRule, IconSetRule
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image
from openpyxl.worksheet.datavalidation import DataValidation

class ExcelGenerator:
    """Advanced Excel generation with professional formatting"""
    
    def __init__(self):
        self.setup_styles()
        self.colors = {
            'primary_blue': '2C57FA',
            'primary_orange': 'FA962C',
            'success_green': '1B5E20',
            'warning_orange': 'F57C00',
            'danger_red': 'C62828',
            'light_gray': 'F8F9FA',
            'medium_gray': '6C757D',
            'dark_gray': '343A40'
        }
    
    def setup_styles(self):
        """Setup reusable styles"""
        # Header style
        self.header_style = NamedStyle(name='header')
        self.header_style.font = Font(bold=True, color='FFFFFF', size=12)
        self.header_style.fill = PatternFill(
            start_color='2C57FA',
            end_color='2C57FA',
            fill_type='solid'
        )
        self.header_style.alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True
        )
        self.header_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title style
        self.title_style = NamedStyle(name='title')
        self.title_style.font = Font(bold=True, size=16, color='2C57FA')
        self.title_style.alignment = Alignment(horizontal='center', vertical='center')
        
        # Subtitle style
        self.subtitle_style = NamedStyle(name='subtitle')
        self.subtitle_style.font = Font(italic=True, size=12, color='6C757D')
        self.subtitle_style.alignment = Alignment(horizontal='center', vertical='center')
        
        # Success style
        self.success_style = NamedStyle(name='success')
        self.success_style.fill = PatternFill(
            start_color='E8F5E9',
            end_color='E8F5E9',
            fill_type='solid'
        )
        self.success_style.font = Font(color='1B5E20')
        
        # Warning style
        self.warning_style = NamedStyle(name='warning')
        self.warning_style.fill = PatternFill(
            start_color='FFF8E1',
            end_color='FFF8E1',
            fill_type='solid'
        )
        self.warning_style.font = Font(color='F57C00')
        
        # Danger style
        self.danger_style = NamedStyle(name='danger')
        self.danger_style.fill = PatternFill(
            start_color='FFEBEE',
            end_color='FFEBEE',
            fill_type='solid'
        )
        self.danger_style.font = Font(color='C62828')
    
    def generate_advanced_excel(self, data: pd.DataFrame) -> bytes:
        """Generate advanced Excel dashboard with multiple sheets"""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Add custom styles to workbook
        self._add_styles_to_workbook(wb)
        
        # Create sheets
        self._create_dashboard_sheet(wb, data)
        self._create_detailed_kpi_sheet(wb, data)
        self._create_performance_matrix_sheet(wb, data)
        self._create_analytics_sheet(wb, data)
        self._create_risk_analysis_sheet(wb, data)
        self._create_timeline_sheet(wb, data)
        self._create_summary_sheet(wb, data)
        self._create_data_sheet(wb, data)
        
        # Add document properties
        self._set_document_properties(wb)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    def generate_simple_excel(self, data: pd.DataFrame) -> bytes:
        """Generate simple Excel export"""
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Main data sheet
            data.to_excel(writer, sheet_name='KPI Data', index=False)
            
            # Summary sheet
            summary = self._create_summary_dataframe(data)
            summary.to_excel(writer, sheet_name='Summary', index=False)
            
            # Format sheets
            workbook = writer.book
            
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                self._apply_basic_formatting(worksheet)
        
        output.seek(0)
        return output.getvalue()
    
    def _create_dashboard_sheet(self, wb: Workbook, data: pd.DataFrame):
        """Create main dashboard sheet"""
        ws = wb.create_sheet('Dashboard')
        
        # Title
        ws.merge_cells('A1:J2')
        ws['A1'] = 'KPI DASHBOARD'
        ws['A1'].style = self.title_style
        
        # Subtitle with date
        ws.merge_cells('A3:J3')
        ws['A3'] = f'Generated on {datetime.now().strftime("%Y-%m-%d %H:%M")}'
        ws['A3'].style = self.subtitle_style
        
        # Key metrics section
        self._add_key_metrics(ws, data, start_row=5)
        
        # Status distribution chart
        self._add_status_chart(ws, data, start_row=12)
        
        # Top KPIs table
        self._add_top_kpis_table(ws, data, start_row=30)
        
        # Apply sheet formatting
        self._format_dashboard_sheet(ws)
    
    def _create_detailed_kpi_sheet(self, wb: Workbook, data: pd.DataFrame):
        """Create detailed KPI listing sheet"""
        ws = wb.create_sheet('KPI Details')
        
        # Headers
        headers = [
            'KPI Name', 'Project', 'Goal', 'Owner', 'Status',
            'Progress', 'Health Score', 'Target', 'Actual', 
            'Completion %', 'Last Updated', 'Risk Level'
        ]
        
        # Map data columns to headers
        column_mapping = {
            'KPI Name': 'kpi_name',
            'Project': 'project',
            'Goal': 'goal',
            'Owner': 'owner',
            'Status': 'status',
            'Progress': 'progress',
            'Health Score': 'health_score',
            'Target': 'target_value',
            'Actual': 'actual_value',
            'Last Updated': 'last_updated',
            'Risk Level': 'risk_level'
        }
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.style = self.header_style
        
        # Write data
        for row_idx, row_data in enumerate(data.iterrows(), 2):
            _, row = row_data
            
            for col_idx, header in enumerate(headers, 1):
                if header == 'Completion %':
                    # Calculate completion percentage
                    target = row.get('target_value', 100)
                    actual = row.get('actual_value', 0)
                    value = (actual / target * 100) if target > 0 else 0
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.number_format = '0.0%'
                else:
                    # Get value from data
                    data_col = column_mapping.get(header)
                    if data_col:
                        value = row.get(data_col, '')
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        
                        # Format based on column type
                        if header == 'Status':
                            self._apply_status_formatting(cell, value)
                        elif header == 'Progress':
                            self._apply_progress_formatting(cell, value)
                        elif header == 'Health Score':
                            cell.number_format = '0.0'
                            self._apply_health_score_formatting(cell, value)
                        elif header in ['Target', 'Actual']:
                            cell.number_format = '#,##0.00'
                        elif header == 'Last Updated':
                            cell.number_format = 'yyyy-mm-dd'
        
        # Add filters
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(data) + 1}"
        
        # Add conditional formatting
        self._add_conditional_formatting(ws, len(data) + 1)
        
        # Create table
        table = Table(
            displayName="KPITable",
            ref=f"A1:{get_column_letter(len(headers))}{len(data) + 1}"
        )
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        ws.add_table(table)
        
        # Adjust column widths
        self._auto_adjust_column_widths(ws)
    
    def _create_performance_matrix_sheet(self, wb: Workbook, data: pd.DataFrame):
        """Create performance matrix sheet"""
        ws = wb.create_sheet('Performance Matrix')
        
        # Title
        ws['A1'] = 'PERFORMANCE MATRIX'
        ws['A1'].style = self.title_style
        ws.merge_cells('A1:F1')
        
        # Create matrix data
        if 'project' in data.columns and 'status' in data.columns:
            matrix = data.pivot_table(
                index='project',
                columns='status',
                values='kpi_name',
                aggfunc='count',
                fill_value=0
            )
            
            # Write matrix to sheet
            start_row = 3
            
            # Headers
            ws.cell(row=start_row, column=1, value='Project').style = self.header_style
            
            col_idx = 2
            for status in ['G', 'Y', 'R']:
                if status in matrix.columns:
                    cell = ws.cell(row=start_row, column=col_idx, value=self._get_status_label(status))
                    cell.style = self.header_style
                    col_idx += 1
            
            ws.cell(row=start_row, column=col_idx, value='Total').style = self.header_style
            ws.cell(row=start_row, column=col_idx + 1, value='Health %').style = self.header_style
            
            # Data rows
            for row_idx, (project, row_data) in enumerate(matrix.iterrows(), start_row + 1):
                ws.cell(row=row_idx, column=1, value=project)
                
                col_idx = 2
                row_total = 0
                green_count = 0
                
                for status in ['G', 'Y', 'R']:
                    if status in matrix.columns:
                        value = row_data[status]
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        
                        if status == 'G':
                            green_count = value
                            cell.style = self.success_style
                        elif status == 'Y':
                            cell.style = self.warning_style
                        elif status == 'R':
                            cell.style = self.danger_style
                        
                        row_total += value
                        col_idx += 1
                
                # Total
                ws.cell(row=row_idx, column=col_idx, value=row_total)
                
                # Health percentage
                health_pct = (green_count / row_total * 100) if row_total > 0 else 0
                cell = ws.cell(row=row_idx, column=col_idx + 1, value=health_pct)
                cell.number_format = '0.0%'
                self._apply_health_score_formatting(cell, health_pct)
            
            # Add chart
            self._add_matrix_chart(ws, matrix, start_row + len(matrix) + 3)
        
        self._auto_adjust_column_widths(ws)
    
    def _create_analytics_sheet(self, wb: Workbook, data: pd.DataFrame):
        """Create analytics sheet with insights"""
        ws = wb.create_sheet('Analytics')
        
        ws['A1'] = 'ANALYTICS & INSIGHTS'
        ws['A1'].style = self.title_style
        ws.merge_cells('A1:H1')
        
        current_row = 3
        
        # Statistical summary
        ws['A' + str(current_row)] = 'Statistical Summary'
        ws['A' + str(current_row)].font = Font(bold=True, size=14)
        current_row += 1
        
        stats = self._calculate_statistics(data)
        for stat_name, stat_value in stats.items():
            ws.cell(row=current_row, column=1, value=stat_name)
            ws.cell(row=current_row, column=2, value=stat_value)
            current_row += 1
        
        current_row += 2
        
        # Insights
        ws['A' + str(current_row)] = 'Key Insights'
        ws['A' + str(current_row)].font = Font(bold=True, size=14)
        current_row += 1
        
        insights = self._generate_insights(data)
        for insight in insights:
            ws.cell(row=current_row, column=1, value='•')
            ws.cell(row=current_row, column=2, value=insight)
            current_row += 1
        
        current_row += 2
        
        # Recommendations
        ws['A' + str(current_row)] = 'Recommendations'
        ws['A' + str(current_row)].font = Font(bold=True, size=14)
        current_row += 1
        
        recommendations = self._generate_recommendations(data)
        for rec in recommendations:
            ws.cell(row=current_row, column=1, value='→')
            ws.cell(row=current_row, column=2, value=rec)
            current_row += 1
        
        self._auto_adjust_column_widths(ws)
    
    def _create_risk_analysis_sheet(self, wb: Workbook, data: pd.DataFrame):
        """Create risk analysis sheet"""
        ws = wb.create_sheet('Risk Analysis')
        
        ws['A1'] = 'RISK ANALYSIS'
        ws['A1'].style = self.title_style
        ws.merge_cells('A1:G1')
        
        # Calculate risk scores
        risk_data = self._calculate_risk_scores(data)
        
        # Headers
        headers = ['KPI Name', 'Project', 'Risk Score', 'Risk Level', 
                  'Status', 'Days Since Update', 'Mitigation']
        
        start_row = 3
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.style = self.header_style
        
        # Sort by risk score
        risk_data_sorted = risk_data.sort_values('risk_score', ascending=False)
        
        # Write risk data
        for row_idx, (_, row) in enumerate(risk_data_sorted.iterrows(), start_row + 1):
            ws.cell(row=row_idx, column=1, value=row.get('kpi_name', ''))
            ws.cell(row=row_idx, column=2, value=row.get('project', ''))
            
            risk_score = row.get('risk_score', 0)
            cell = ws.cell(row=row_idx, column=3, value=risk_score)
            cell.number_format = '0.0'
            
            # Risk level
            risk_level = self._get_risk_level(risk_score)
            cell = ws.cell(row=row_idx, column=4, value=risk_level)
            self._apply_risk_formatting(cell, risk_level)
            
            # Status
            status = row.get('status', 'Y')
            cell = ws.cell(row=row_idx, column=5, value=status)
            self._apply_status_formatting(cell, status)
            
            # Days since update
            ws.cell(row=row_idx, column=6, value=row.get('days_since_update', 0))
            
            # Mitigation suggestion
            ws.cell(row=row_idx, column=7, value=self._get_mitigation_suggestion(row))
        
        # Add risk distribution chart
        self._add_risk_chart(ws, risk_data_sorted, start_row + len(risk_data_sorted) + 3)
        
        self._auto_adjust_column_widths(ws)
    
    def _create_timeline_sheet(self, wb: Workbook, data: pd.DataFrame):
        """Create timeline/Gantt sheet"""
        ws = wb.create_sheet('Timeline')
        
        ws['A1'] = 'PROJECT TIMELINE'
        ws['A1'].style = self.title_style
        ws.merge_cells('A1:H1')
        
        if 'last_updated' in data.columns:
            # Group by project and create timeline view
            timeline_data = []
            
            for project in data['project'].unique():
                project_data = data[data['project'] == project]
                
                timeline_data.append({
                    'Project': project,
                    'KPIs': len(project_data),
                    'Start Date': project_data['last_updated'].min(),
                    'Last Update': project_data['last_updated'].max(),
                    'Completion %': (project_data['actual_value'].sum() / 
                                   project_data['target_value'].sum() * 100 
                                   if project_data['target_value'].sum() > 0 else 0),
                    'Status': self._get_overall_status(project_data)
                })
            
            timeline_df = pd.DataFrame(timeline_data)
            
            # Write timeline data
            start_row = 3
            headers = timeline_df.columns.tolist()
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col, value=header)
                cell.style = self.header_style
            
            for row_idx, (_, row) in enumerate(timeline_df.iterrows(), start_row + 1):
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    
                    if headers[col_idx - 1] == 'Completion %':
                        cell.number_format = '0.0%'
                    elif headers[col_idx - 1] in ['Start Date', 'Last Update']:
                        cell.number_format = 'yyyy-mm-dd'
                    elif headers[col_idx - 1] == 'Status':
                        self._apply_status_formatting(cell, value)
            
            # Add timeline chart
            self._add_timeline_chart(ws, timeline_df, start_row + len(timeline_df) + 3)
        
        self._auto_adjust_column_widths(ws)
    
    def _create_summary_sheet(self, wb: Workbook, data: pd.DataFrame):
        """Create executive summary sheet"""
        ws = wb.create_sheet('Summary')
        
        ws['A1'] = 'EXECUTIVE SUMMARY'
        ws['A1'].style = self.title_style
        ws.merge_cells('A1:E1')
        
        summary_data = self._create_summary_dataframe(data)
        
        # Write summary data
        start_row = 3
        
        for row_idx, (metric, value) in enumerate(summary_data.items(), start_row):
            ws.cell(row=row_idx, column=1, value=metric).font = Font(bold=True)
            ws.cell(row=row_idx, column=2, value=value)
        
        # Add summary charts
        self._add_summary_charts(ws, data, start_row + len(summary_data) + 3)
        
        self._auto_adjust_column_widths(ws)
    
    def _create_data_sheet(self, wb: Workbook, data: pd.DataFrame):
        """Create raw data sheet with validation"""
        ws = wb.create_sheet('Raw Data')
        
        # Write dataframe to sheet
        for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                if r_idx == 1:
                    cell.style = self.header_style
        
        # Add data validation
        self._add_data_validation(ws, len(data) + 1)
        
        # Freeze panes
        ws.freeze_panes = 'A2'
        
        self._auto_adjust_column_widths(ws)
    
    # Helper methods
    def _add_styles_to_workbook(self, wb: Workbook):
        """Add custom styles to workbook"""
        styles = [
            self.header_style, self.title_style, self.subtitle_style,
            self.success_style, self.warning_style, self.danger_style
        ]
        
        for style in styles:
            try:
                wb.add_named_style(style)
            except:
                pass  # Style already exists
    
    def _add_key_metrics(self, ws, data: pd.DataFrame, start_row: int):
        """Add key metrics section to dashboard"""
        metrics = self._calculate_key_metrics(data)
        
        # Metric headers
        metric_headers = ['Total KPIs', 'On Track', 'At Risk', 'Avg Health', 'Avg Progress']
        metric_values = [
            metrics['total'],
            metrics['on_track'],
            metrics['at_risk'],
            f"{metrics['avg_health']:.0f}%",
            f"{metrics['avg_progress']:.1f}/5"
        ]
        
        # Write metrics
        for col, (header, value) in enumerate(zip(metric_headers, metric_values), 1):
            # Header
            header_cell = ws.cell(row=start_row, column=col * 2 - 1, value=header)
            header_cell.font = Font(bold=True, size=10, color='6C757D')
            
            # Value
            value_cell = ws.cell(row=start_row + 1, column=col * 2 - 1, value=value)
            value_cell.font = Font(bold=True, size=16, color='2C57FA')
            
            # Format based on metric type
            if header == 'On Track':
                value_cell.font = Font(bold=True, size=16, color='1B5E20')
            elif header == 'At Risk':
                value_cell.font = Font(bold=True, size=16, color='C62828')
    
    def _add_status_chart(self, ws, data: pd.DataFrame, start_row: int):
        """Add status distribution chart"""
        if 'status' not in data.columns:
            return
        
        # Calculate status counts
        status_counts = data['status'].value_counts()
        
        # Write data for chart
        ws.cell(row=start_row, column=1, value='Status')
        ws.cell(row=start_row, column=2, value='Count')
        
        row = start_row + 1
        for status in ['G', 'Y', 'R']:
            if status in status_counts.index:
                ws.cell(row=row, column=1, value=self._get_status_label(status))
                ws.cell(row=row, column=2, value=status_counts[status])
                row += 1
        
        # Create pie chart
        chart = PieChart()
        chart.title = "KPI Status Distribution"
        chart.style = 10
        
        data_ref = Reference(ws, min_col=2, min_row=start_row, max_row=row - 1, max_col=2)
        labels = Reference(ws, min_col=1, min_row=start_row + 1, max_row=row - 1)
        
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(labels)
        
        ws.add_chart(chart, f"E{start_row}")
    
    def _add_top_kpis_table(self, ws, data: pd.DataFrame, start_row: int):
        """Add top performing KPIs table"""
        ws.cell(row=start_row, column=1, value='TOP PERFORMING KPIs').font = Font(bold=True, size=14)
        
        # Get top KPIs by health score
        if 'health_score' in data.columns:
            top_kpis = data.nlargest(5, 'health_score')[['kpi_name', 'project', 'health_score', 'status']]
            
            # Headers
            headers = ['KPI Name', 'Project', 'Health Score', 'Status']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=start_row + 2, column=col, value=header)
                cell.style = self.header_style
            
            # Data
            for row_idx, (_, row) in enumerate(top_kpis.iterrows(), start_row + 3):
                ws.cell(row=row_idx, column=1, value=row['kpi_name'])
                ws.cell(row=row_idx, column=2, value=row['project'])
                
                health_cell = ws.cell(row=row_idx, column=3, value=row['health_score'])
                health_cell.number_format = '0.0'
                
                status_cell = ws.cell(row=row_idx, column=4, value=row['status'])
                self._apply_status_formatting(status_cell, row['status'])
    
    def _format_dashboard_sheet(self, ws):
        """Apply formatting to dashboard sheet"""
        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20
        
        # Set row heights
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 25
    
    def _apply_status_formatting(self, cell, status):
        """Apply formatting based on status"""
        if status == 'G':
            cell.style = self.success_style
            cell.value = '🟢 On Track'
        elif status == 'Y':
            cell.style = self.warning_style
            cell.value = '🟡 Needs Attention'
        elif status == 'R':
            cell.style = self.danger_style
            cell.value = '🔴 At Risk'
    
    def _apply_progress_formatting(self, cell, progress):
        """Apply formatting based on progress level"""
        try:
            progress_val = int(progress)
            
            if progress_val >= 4:
                cell.font = Font(color='1B5E20')
            elif progress_val >= 3:
                cell.font = Font(color='F57C00')
            else:
                cell.font = Font(color='C62828')
            
            # Add progress bar using conditional formatting
            cell.value = f"{progress_val}/5"
            
        except:
            cell.value = progress
    
    def _apply_health_score_formatting(self, cell, score):
        """Apply formatting based on health score"""
        try:
            score_val = float(score)
            
            if score_val >= 80:
                cell.font = Font(color='1B5E20', bold=True)
            elif score_val >= 60:
                cell.font = Font(color='F57C00')
            else:
                cell.font = Font(color='C62828', bold=True)
                
        except:
            pass
    
    def _apply_risk_formatting(self, cell, risk_level):
        """Apply formatting based on risk level"""
        if risk_level == 'High':
            cell.style = self.danger_style
        elif risk_level == 'Medium':
            cell.style = self.warning_style
        elif risk_level == 'Low':
            cell.style = self.success_style
    
    def _add_conditional_formatting(self, ws, max_row: int):
        """Add conditional formatting rules"""
        # Color scale for health scores
        if 'G' in [cell.value for cell in ws[1]]:
            health_col = None
            for cell in ws[1]:
                if cell.value == 'Health Score':
                    health_col = cell.column_letter
                    break
            
            if health_col:
                ws.conditional_formatting.add(
                    f'{health_col}2:{health_col}{max_row}',
                    ColorScaleRule(
                        start_type='min',
                        start_color='FF0000',
                        mid_type='percentile',
                        mid_value=50,
                        mid_color='FFFF00',
                        end_type='max',
                        end_color='00FF00'
                    )
                )
        
        # Data bars for completion percentage
        for cell in ws[1]:
            if cell.value == 'Completion %':
                comp_col = cell.column_letter
                ws.conditional_formatting.add(
                    f'{comp_col}2:{comp_col}{max_row}',
                    DataBarRule(
                        start_type='num',
                        start_value=0,
                        end_type='num',
                        end_value=100,
                        color='2C57FA'
                    )
                )
                break
    
    def _add_data_validation(self, ws, max_row: int):
        """Add data validation to cells"""
        # Status validation
        status_validation = DataValidation(
            type="list",
            formula1='"G,Y,R"',
            allow_blank=False,
            showDropDown=True,
            errorTitle='Invalid Status',
            error='Please select G (Green), Y (Yellow), or R (Red)'
        )
        
        # Progress validation
        progress_validation = DataValidation(
            type="whole",
            operator="between",
            formula1=1,
            formula2=5,
            allow_blank=False,
            errorTitle='Invalid Progress',
            error='Progress must be between 1 and 5'
        )
        
        # Apply validations
        for cell in ws[1]:
            if cell.value == 'Status':
                status_col = cell.column_letter
                status_validation.add(f'{status_col}2:{status_col}{max_row}')
                ws.add_data_validation(status_validation)
            elif cell.value == 'Progress':
                progress_col = cell.column_letter
                progress_validation.add(f'{progress_col}2:{progress_col}{max_row}')
                ws.add_data_validation(progress_validation)
    
    def _auto_adjust_column_widths(self, ws):
        """Auto-adjust column widths based on content"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _set_document_properties(self, wb: Workbook):
        """Set document properties"""
        wb.properties.title = "KPI Dashboard"
        wb.properties.subject = "Performance Tracking"
        wb.properties.creator = "Enhanced KPI System"
        wb.properties.description = "Comprehensive KPI tracking and analytics dashboard"
        wb.properties.created = datetime.now()
        wb.properties.modified = datetime.now()
    
    # Calculation methods
    def _calculate_key_metrics(self, data: pd.DataFrame) -> Dict:
        """Calculate key metrics from data"""
        metrics = {
            'total': len(data),
            'on_track': 0,
            'at_risk': 0,
            'avg_health': 0,
            'avg_progress': 0
        }
        
        if 'status' in data.columns:
            metrics['on_track'] = (data['status'] == 'G').sum()
            metrics['at_risk'] = (data['status'] == 'R').sum()
        
        if 'health_score' in data.columns:
            metrics['avg_health'] = data['health_score'].mean()
        
        if 'progress' in data.columns:
            metrics['avg_progress'] = data['progress'].mean()
        
        return metrics
    
    def _calculate_statistics(self, data: pd.DataFrame) -> Dict:
        """Calculate statistical summary"""
        stats = {}
        
        numeric_columns = data.select_dtypes(include=[np.number]).columns
        
        for col in numeric_columns:
            if col in ['health_score', 'progress', 'target_value', 'actual_value']:
                stats[f'{col}_mean'] = data[col].mean()
                stats[f'{col}_median'] = data[col].median()
                stats[f'{col}_std'] = data[col].std()
                stats[f'{col}_min'] = data[col].min()
                stats[f'{col}_max'] = data[col].max()
        
        return stats
    
    def _calculate_risk_scores(self, data: pd.DataFrame) -> pd.DataFrame:
        """Calculate risk scores for KPIs"""
        risk_data = data.copy()
        
        risk_data['risk_score'] = 0
        
        # Status risk
        if 'status' in risk_data.columns:
            risk_data.loc[risk_data['status'] == 'R', 'risk_score'] += 40
            risk_data.loc[risk_data['status'] == 'Y', 'risk_score'] += 20
        
        # Progress risk
        if 'progress' in risk_data.columns:
            risk_data.loc[risk_data['progress'] <= 2, 'risk_score'] += 30
            risk_data.loc[risk_data['progress'] == 3, 'risk_score'] += 15
        
        # Health score risk
        if 'health_score' in risk_data.columns:
            risk_data.loc[risk_data['health_score'] < 50, 'risk_score'] += 20
            risk_data.loc[(risk_data['health_score'] >= 50) & (risk_data['health_score'] < 70), 'risk_score'] += 10
        
        # Days since update
        if 'last_updated' in risk_data.columns:
            risk_data['days_since_update'] = (datetime.now() - pd.to_datetime(risk_data['last_updated'])).dt.days
            risk_data.loc[risk_data['days_since_update'] > 30, 'risk_score'] += 10
            risk_data.loc[risk_data['days_since_update'] > 14, 'risk_score'] += 5
        
        return risk_data
    
    def _create_summary_dataframe(self, data: pd.DataFrame) -> Dict:
        """Create summary statistics"""
        summary = {
            'Total KPIs': len(data),
            'Unique Projects': data['project'].nunique() if 'project' in data.columns else 0,
            'Unique Owners': data['owner'].nunique() if 'owner' in data.columns else 0
        }
        
        if 'status' in data.columns:
            status_counts = data['status'].value_counts()
            summary['On Track (Green)'] = status_counts.get('G', 0)
            summary['Needs Attention (Yellow)'] = status_counts.get('Y', 0)
            summary['At Risk (Red)'] = status_counts.get('R', 0)
        
        if 'health_score' in data.columns:
            summary['Average Health Score'] = f"{data['health_score'].mean():.1f}%"
            summary['Highest Health Score'] = f"{data['health_score'].max():.1f}%"
            summary['Lowest Health Score'] = f"{data['health_score'].min():.1f}%"
        
        if 'progress' in data.columns:
            summary['Average Progress'] = f"{data['progress'].mean():.1f}/5"
        
        if 'target_value' in data.columns and 'actual_value' in data.columns:
            overall_completion = (data['actual_value'].sum() / data['target_value'].sum() * 100 
                                if data['target_value'].sum() > 0 else 0)
            summary['Overall Completion'] = f"{overall_completion:.1f}%"
        
        return summary
    
    # Label and formatting helpers
    def _get_status_label(self, status: str) -> str:
        """Get label for status code"""
        labels = {
            'G': 'On Track',
            'Y': 'Needs Attention',
            'R': 'At Risk'
        }
        return labels.get(status, status)
    
    def _get_risk_level(self, risk_score: float) -> str:
        """Get risk level from score"""
        if risk_score >= 70:
            return 'High'
        elif risk_score >= 40:
            return 'Medium'
        else:
            return 'Low'
    
    def _get_mitigation_suggestion(self, row: pd.Series) -> str:
        """Generate mitigation suggestion based on risk factors"""
        suggestions = []
        
        if row.get('status') == 'R':
            suggestions.append("Immediate intervention required")
        
        if row.get('progress', 5) <= 2:
            suggestions.append("Accelerate progress activities")
        
        if row.get('days_since_update', 0) > 14:
            suggestions.append("Update KPI status")
        
        if row.get('health_score', 100) < 50:
            suggestions.append("Review and revise targets")
        
        return '; '.join(suggestions) if suggestions else 'Monitor closely'
    
    def _get_overall_status(self, project_data: pd.DataFrame) -> str:
        """Get overall status for a project"""
        if 'status' not in project_data.columns:
            return 'Unknown'
        
        status_counts = project_data['status'].value_counts()
        
        # If any are red, overall is red
        if status_counts.get('R', 0) > 0:
            return 'R'
        # If majority are yellow, overall is yellow
        elif status_counts.get('Y', 0) > status_counts.get('G', 0):
            return 'Y'
        else:
            return 'G'
    
    # Insight generation
    def _generate_insights(self, data: pd.DataFrame) -> List[str]:
        """Generate insights from data"""
        insights = []
        
        if 'status' in data.columns:
            at_risk_pct = (data['status'] == 'R').mean() * 100
            if at_risk_pct > 30:
                insights.append(f"{at_risk_pct:.0f}% of KPIs are at risk - immediate action needed")
            
            on_track_pct = (data['status'] == 'G').mean() * 100
            if on_track_pct > 70:
                insights.append(f"Strong performance with {on_track_pct:.0f}% of KPIs on track")
        
        if 'health_score' in data.columns:
            low_health = data[data['health_score'] < 50]
            if len(low_health) > 0:
                insights.append(f"{len(low_health)} KPIs have critically low health scores")
        
        if 'last_updated' in data.columns:
            stale_kpis = data[(datetime.now() - pd.to_datetime(data['last_updated'])).dt.days > 14]
            if len(stale_kpis) > 0:
                insights.append(f"{len(stale_kpis)} KPIs haven't been updated in 14+ days")
        
        if 'progress' in data.columns:
            low_progress = data[data['progress'] <= 2]
            if len(low_progress) > 0:
                insights.append(f"{len(low_progress)} KPIs showing limited progress")
        
        return insights if insights else ["All KPIs are within normal parameters"]
    
    def _generate_recommendations(self, data: pd.DataFrame) -> List[str]:
        """Generate recommendations from data"""
        recommendations = []
        
        # Check for at-risk KPIs
        if 'status' in data.columns:
            at_risk = data[data['status'] == 'R']
            if len(at_risk) > 0:
                recommendations.append(f"Focus immediate attention on {len(at_risk)} at-risk KPIs")
        
        # Check for stale data
        if 'last_updated' in data.columns:
            stale = data[(datetime.now() - pd.to_datetime(data['last_updated'])).dt.days > 7]
            if len(stale) > 0:
                recommendations.append(f"Update {len(stale)} KPIs that are more than a week old")
        
        # Check for low performers
        if 'health_score' in data.columns:
            low_performers = data[data['health_score'] < 60]
            if len(low_performers) > 0:
                recommendations.append(f"Develop improvement plans for {len(low_performers)} low-performing KPIs")
        
        # Check for resource allocation
        if 'owner' in data.columns:
            owner_load = data['owner'].value_counts()
            overloaded = owner_load[owner_load > 10]
            if len(overloaded) > 0:
                recommendations.append(f"Consider redistributing KPIs from overloaded owners: {', '.join(overloaded.index)}")
        
        return recommendations if recommendations else ["Continue current monitoring and support activities"]
    
    # Chart creation methods
    def _add_matrix_chart(self, ws, matrix_data, start_row: int):
        """Add matrix visualization chart"""
        chart = BarChart()
        chart.title = "KPIs by Project and Status"
        chart.style = 10
        chart.type = "col"
        chart.grouping = "stacked"
        
        # Add to worksheet
        ws.add_chart(chart, f"A{start_row}")
    
    def _add_risk_chart(self, ws, risk_data, start_row: int):
        """Add risk distribution chart"""
        chart = BarChart()
        chart.title = "Risk Distribution"
        chart.style = 11
        
        ws.add_chart(chart, f"A{start_row}")
    
    def _add_timeline_chart(self, ws, timeline_data, start_row: int):
        """Add timeline/Gantt chart"""
        chart = LineChart()
        chart.title = "Project Timeline"
        chart.style = 12
        
        ws.add_chart(chart, f"A{start_row}")
    
    def _add_summary_charts(self, ws, data, start_row: int):
        """Add summary charts"""
        # Create multiple small charts for summary
        if 'status' in data.columns:
            # Status pie chart
            pie = PieChart()
            pie.title = "Status Distribution"
            ws.add_chart(pie, f"D{start_row}")
        
        if 'progress' in data.columns:
            # Progress bar chart
            bar = BarChart()
            bar.title = "Progress Distribution"
            ws.add_chart(bar, f"H{start_row}")